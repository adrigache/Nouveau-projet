"""Final details: CAPEX breakdown, Overview ratios, SPV_6 engine vs Excel CFADS, Computations INDEX."""
from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter

WB = "model_real.xlsx"


def main():
    wb = openpyxl.load_workbook(WB, data_only=False)
    h1 = wb["Holdco_1"]

    print("=== Holdco_1 CAPEX block labels+formulas R391-430 col R ===")
    for r in range(391, 432):
        lab = h1.cell(r, 3).value
        v = h1.cell(r, 18).value
        if lab or v:
            print(f"R{r}: {lab} || {v}")

    print("\n=== Holdco_1 detention helpers R229-262 formulas R ===")
    for r in range(229, 263):
        lab = h1.cell(r, 3).value
        v = h1.cell(r, 18).value
        if lab or v:
            print(f"R{r}: {lab} || {v}")

    print("\n=== Holdco_1 r1155 / share capital detention source ===")
    for r in (1155, 1156, 1162, 1163, 1164):
        print(f"R{r} C3={h1.cell(r,3).value} R18={h1.cell(r,18).value}")

    print("\n=== Overview ratios section formulas R130-160 ===")
    ov = wb["Overview"]
    for r in range(130, 165):
        labs = [(c, ov.cell(r, c).value) for c in range(1, 14) if ov.cell(r, c).value is not None]
        q = ov.cell(r, 17).value
        rr = ov.cell(r, 18).value
        if labs or q or rr:
            print(f"R{r}: {labs}")
            print(f"     Q={q} R={rr}")

    print("\n=== Overview B/S check section ===")
    for r in range(95, 130):
        labs = [(c, ov.cell(r, c).value) for c in range(1, 14) if ov.cell(r, c).value is not None]
        if labs:
            print(f"R{r}: {labs}")
            print(f"     Q={ov.cell(r,17).value} R={ov.cell(r,18).value}")

    print("\n=== Computations G4 name formula / INDEX vehicle sheets ===")
    comp = wb["Computations"]
    print("G4:", comp.cell(4, 7).value)
    # search for INDEX/CHOOSE/INDIRECT linking sheets
    for r in range(1, 50):
        for c in range(1, 20):
            v = comp.cell(r, c).value
            if isinstance(v, str) and any(k in v for k in ["Holdco", "SPV_", "INDIRECT", "CHOOSE", "INDEX"]):
                print(f"Computations R{r}C{c}: {v[:200]}")

    # named ranges related to ref_Vehicle - check defined names
    print("\n=== Defined names containing Vehicle/ref ===")
    for dn in wb.defined_names.values():
        name = dn.name
        if any(k in name.lower() for k in ["vehicle", "ref_", "calage", "master"]):
            print(f"  {name} = {dn.attr_text}")

    print("\n=== SPV_6 Flag distribution + EBITDA/EBIT/CIT path ===")
    s6 = wb["SPV_6"]
    for r in [139, 141, 142, 150, 152, 153, 325, 329, 343, 344, 345, 429, 430, 451]:
        print(f"R{r}: {s6.cell(r,3).value}")
        print(f"  R18={s6.cell(r,18).value}")
        print(f"  R19={s6.cell(r,19).value}")

    print("\n=== ass_Vehicle gearing / max debt / LTV Holdco ===")
    veh = wb["ass_Vehicle"]
    for r in range(210, 250):
        lab = None
        for c in range(2, 12):
            v = veh.cell(r, c).value
            if isinstance(v, str) and v.strip():
                lab = v.strip()
                break
        if lab:
            print(f"R{r}: {lab[:80]} | H1={veh.cell(r,14).value} H2={veh.cell(r,15).value}")

    wb.close()

    wb2 = openpyxl.load_workbook(WB, data_only=True)
    h1 = wb2["Holdco_1"]
    print("\n=== Holdco_1 YE2026 CAPEX/DA values ===")
    for r in range(391, 432):
        lab = h1.cell(r, 3).value
        v = h1.cell(r, 18).value
        if v not in (None, 0, 0.0) or (isinstance(lab, str) and lab):
            if v not in (None, 0, 0.0):
                print(f"R{r}: {lab} = {v}")

    print("\n=== Holdco_1 YE2026 Assets path ===")
    for r in [427, 428, 429, 430, 158]:
        print(f"R{r}: {h1.cell(r,3).value} = {h1.cell(r,18).value}")

    print("\n=== Holdco_1 YE2025 vs 2026 senior / SHL / RE / proceeds ===")
    for r, name in [(158,'Assets'),(159,'SC det'),(160,'SHL SPVs'),(164,'Proceeds'),
                    (169,'SC'),(170,'LR'),(171,'RE'),(173,'SHL'),(175,'Senior'),
                    (177,'EBL'),(180,'CIT pay'),(185,'TL'),(186,'Check')]:
        print(f"{name}: 2025(Q)={h1.cell(r,17).value} 2026(R)={h1.cell(r,18).value}")

    print("\n=== Overview ratios DATA for current entity ===")
    ov = wb2["Overview"]
    print(f"Selected name={ov.cell(4,7).value} ref={ov.cell(3,7).value}")
    for r in range(100, 165):
        lab = ov.cell(r, 3).value
        vals = [ov.cell(r, c).value for c in range(17, 25)]
        nonzero = [v for v in vals if isinstance(v, (int, float)) and abs(v) > 1e-9]
        if isinstance(lab, str) and (nonzero or any(k in lab.lower() for k in [
            "dscr", "llcr", "plcr", "irr", "npv", "gear", "check", "ratio", "hdscr"
        ])):
            print(f"R{r}: {lab} | {vals[:6]}")

    print("\n=== Overview Check B/S ===")
    for r in range(100, 125):
        lab = ov.cell(r, 3).value
        if isinstance(lab, str) and ("check" in lab.lower() or "b/s" in lab.lower() or
                                      "asset" in lab.lower() or "liab" in lab.lower() or
                                      "total" in lab.lower()):
            print(f"R{r}: {lab} | 2026={ov.cell(r,18).value} formula-less val")

    # find Check B/S overview specifically
    for r in range(1, 293):
        for c in range(1, 15):
            v = ov.cell(r, c).value
            if isinstance(v, str) and "check b" in v.lower():
                print(f"FOUND {r},{c}: {v} nearby={[ov.cell(r,cc).value for cc in range(1,15)]}")
                print(f"  values: {[ov.cell(r,cc).value for cc in range(17,25)]}")

    print("\n=== SPV_6 P&L/CIT/DA data 2026/2027 ===")
    s6 = wb2["SPV_6"]
    for r in [139, 141, 142, 150, 152, 153, 325, 329, 343, 344, 345, 429, 430, 451,
              59, 67, 70, 91, 111, 125]:
        print(f"R{r}: {s6.cell(r,3).value} | 26={s6.cell(r,18).value} | 27={s6.cell(r,19).value}")

    wb2.close()

    # Engine comparison
    print("\n=== ENGINE SPV_6 cfads ===")
    import sys
    sys.path.insert(0, "src")
    from pathlib import Path
    from mca_model.plumbing import build
    from mca_model.model import financing as fin
    m = build.load(Path("test/assets/real_model.toml"), debug=False)
    fin.clear_cache()
    spv = m.get_object("SPV_6")
    res = fin.run_spv_financing(m, spv)
    for y in (2025, 2026, 2027, 2028):
        print(f"  {y}: cfads={res.cfads.get(y)} ebitda={res.ebitda.get(y)} ebit={res.ebit.get(y)} "
              f"cit_paid={res.cit_paid.get(y)} cit_due={res.cit_due.get(y)} net={res.net_result.get(y)} "
              f"re={res.retained_earnings_eop.get(y)} proceeds_proxy=N/A")


if __name__ == "__main__":
    main()
