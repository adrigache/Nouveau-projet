"""Focused extraction: Holdco B/S formulas, senior params, Overview, SPV_6 CFADS."""
from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter

WB = "model_real.xlsx"


def formula_or_val(ws, r, c):
    return ws.cell(r, c).value


def dump_row_formulas(ws, rows, cols, title):
    print(f"\n### {title}")
    for r in rows:
        lab = ws.cell(r, 3).value
        print(f"\nR{r}: {lab}")
        for c in cols:
            v = ws.cell(r, c).value
            if v is not None:
                print(f"  {get_column_letter(c)}: {v}")


def main():
    wb = openpyxl.load_workbook(WB, data_only=False)

    # ---- Holdco_1 B/S formulas for col R (2026 = col 18) ----
    h1 = wb["Holdco_1"]
    # confirm year row
    print("Holdco_1 year headers row11:", [(c, h1.cell(11, c).value) for c in range(16, 25)])
    print("Holdco_1 year headers row3:", [(c, h1.cell(3, c).value) for c in range(1, 20)])

    bs_rows = list(range(155, 190))
    dump_row_formulas(h1, bs_rows, [8, 11, 17, 18, 19], "Holdco_1 B/S formulas (Q/R/S + checks)")

    # CF waterfall key rows
    cf_rows = [59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72,
               82, 83, 84, 85, 90, 91, 93, 94, 95, 96, 106, 108, 113, 123, 124, 125, 126]
    dump_row_formulas(h1, cf_rows, [18], "Holdco_1 CF waterfall R2026 formulas")

    # P&L key
    pnl_rows = [129, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142,
                143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153]
    dump_row_formulas(h1, pnl_rows, [18], "Holdco_1 P&L R2026 formulas")

    # CAPEX / D&A
    capex_rows = [28, 29, 53, 54, 55, 391, 403, 411, 418, 426, 427, 428, 429, 430, 431]
    dump_row_formulas(h1, capex_rows, [18], "Holdco_1 CAPEX/DA R2026")

    # Equity / senior schedule headers
    eq_rows = [448, 454, 455, 456, 459, 461, 466, 486, 501, 507, 509, 530, 533, 538,
               543, 632, 633, 634, 635, 636, 637, 638, 639, 640, 641, 642]
    dump_row_formulas(h1, eq_rows, [18], "Holdco_1 Equity/Senior R2026")

    # Senior facility more rows
    print("\n### Senior facility block labels R543-700")
    for r in range(543, 701):
        lab = h1.cell(r, 3).value
        if lab:
            v = h1.cell(r, 18).value
            print(f"R{r}: {lab} || R18={v}")

    # Flags top
    print("\n### Holdco_1 flags / refs R1-50")
    for r in range(1, 51):
        labs = [(c, h1.cell(r, c).value) for c in range(1, 16) if h1.cell(r, c).value is not None]
        if labs:
            print(f"R{r}: {labs}")

    # ---- ass_Vehicle: print ALL rows that differ Holdco vs SPV for financing ----
    veh = wb["ass_Vehicle"]
    print("\n### ass_Vehicle ALL labeled rows 1-480 with Holdco_1 / Holdco_2 / SPV_1 values")
    names = {c: veh.cell(2, c).value for c in range(13, 21)}
    print("cols:", names)
    for r in range(1, 481):
        lab = None
        for c in range(2, 12):
            v = veh.cell(r, c).value
            if isinstance(v, str) and v.strip():
                lab = v.strip()
                break
        if not lab:
            continue
        vals = {names[c]: veh.cell(r, c).value for c in range(13, 21)}
        # skip if all empty/zero/same None
        nonempty = [v for v in vals.values() if v not in (None, "", 0, 0.0, False)]
        # always print if financing-ish OR any holdco nonzero
        low = lab.lower()
        interesting = any(k in low for k in [
            "senior", "debt", "facility", "margin", "interest", "draw", "repay",
            "tenor", "amort", "dsrf", "junior", "crowd", "agency", "upfront",
            "commitment", "gear", "ltv", "funding", "financ", "swap", "euribor",
            "base rate", "spread", "hedge", "sculpt", "target dscr", "sizing",
            "capex", "vehicle", "perimeter", "detention", "held", "holding",
            "share capital", "shl", "cash pool", "fee", "ref", "consolidat",
            "maturity", "financial close", "fc ", "amount", "notional",
        ])
        h_vals = [vals.get("Holdco_1"), vals.get("Holdco_2")]
        if interesting or any(v not in (None, "", 0, 0.0, False) for v in h_vals):
            # compact
            show = {k: vals[k] for k in ("Holdco_1", "Holdco_2", "SPV_1", "Topco") if k in vals}
            print(f"R{r:3d}: {lab[:70]:70s} {show}")

    # ---- ass_Model financing ----
    am = wb["ass_Model"]
    print("\n### ass_Model labeled rows (finance-related) with M/N values")
    for r in range(1, 250):
        labs = []
        for c in range(2, 13):
            v = am.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labs.append((c, v.strip()))
        if not labs:
            continue
        lab = " | ".join(x[1] for x in labs)
        low = lab.lower()
        if any(k in low for k in [
            "senior", "debt", "facility", "margin", "interest", "dsrf", "junior",
            "agency", "upfront", "commitment", "gear", "ltv", "financ", "tenor",
            "amort", "swap", "euribor", "rate", "dscr", "sculpt", "hedge",
            "accuracy", "check", "cit", "atad", "thin",
        ]):
            vals = [(get_column_letter(c), am.cell(r, c).value)
                    for c in range(12, 25) if am.cell(r, c).value is not None]
            print(f"R{r:3d}: {lab[:100]}")
            print(f"       {vals}")

    # ---- Overview: entity selection + ratios section ----
    ov = wb["Overview"]
    print("\n### Overview structure (formulas)")
    for r in range(1, 293):
        labs = [(c, ov.cell(r, c).value) for c in range(1, 16) if ov.cell(r, c).value is not None]
        if labs:
            # also sample Q/R
            extra = [(get_column_letter(c), ov.cell(r, c).value)
                     for c in (17, 18) if ov.cell(r, c).value is not None]
            print(f"R{r}: {labs}")
            if extra:
                print(f"     sample: {extra}")

    # ---- Computations sheet: how entity selected ----
    comp = wb["Computations"]
    print("\n### Computations entity routing (rows 1-30 + named refs)")
    for r in range(1, 30):
        labs = [(c, comp.cell(r, c).value) for c in range(1, 20) if comp.cell(r, c).value is not None]
        if labs:
            print(f"R{r}: {labs}")

    # Dashboard ref_Vehicle
    dash = wb["Dashboard"]
    print("\n### Dashboard vehicle selection (search)")
    for r in range(1, 200):
        for c in range(1, 20):
            v = dash.cell(r, c).value
            if isinstance(v, str) and any(k in v.lower() for k in [
                "vehicle", "holdco", "spv", "ref_", "select", "calage"
            ]):
                print(f"Dashboard R{r}C{c}: {v} | neighbors: "
                      f"{[(cc, dash.cell(r, cc).value) for cc in range(max(1,c-2), c+5) if dash.cell(r, cc).value is not None]}")

    # SPV_6 CFADS path formulas
    spv6 = wb["SPV_6"]
    print("\n### SPV_6 CF / equity waterfall key formulas 2026(R)/2027(S)")
    key_rows = list(range(59, 130)) + list(range(448, 540))
    for r in key_rows:
        lab = spv6.cell(r, 3).value
        if not lab:
            continue
        v18 = spv6.cell(r, 18).value
        v19 = spv6.cell(r, 19).value
        if v18 is not None or v19 is not None:
            print(f"R{r}: {lab}")
            print(f"  R18={v18}")
            print(f"  R19={v19}")

    wb.close()

    # data_only values
    print("\n\n########## DATA_ONLY ##########")
    wb2 = openpyxl.load_workbook(WB, data_only=True)
    h1 = wb2["Holdco_1"]
    print("\n### Holdco_1 YE2026 B/S (col R=18)")
    for r in range(155, 190):
        lab = h1.cell(r, 3).value
        v = h1.cell(r, 18).value
        print(f"R{r}: {lab} = {v}")

    print("\n### Holdco_1 YE2026 key CF/PnL/Equity")
    for r in [59, 66, 72, 82, 83, 91, 94, 125, 126, 139, 142, 144, 151,
              391, 411, 418, 426, 430, 456, 459, 466, 507, 524, 530, 538,
              634, 635, 636, 637, 638, 639, 640]:
        lab = h1.cell(r, 3).value
        print(f"R{r}: {lab} = {h1.cell(r, 18).value}")

    print("\n### Holdco_2 YE2026 B/S")
    h2 = wb2["Holdco_2"]
    for r in range(155, 190):
        lab = h2.cell(r, 3).value
        v = h2.cell(r, 18).value
        print(f"R{r}: {lab} = {v}")

    print("\n### SPV_6 CFADS path values 2026/2027")
    s6 = wb2["SPV_6"]
    for r in list(range(59, 130)) + [448, 451, 456, 490, 491, 496, 507, 510, 524, 530, 536, 538]:
        lab = s6.cell(r, 3).value
        if lab:
            print(f"R{r}: {lab} | 2026={s6.cell(r, 18).value} | 2027={s6.cell(r, 19).value}")

    print("\n### Overview ratios data (selected entity)")
    ov = wb2["Overview"]
    print(f"ref name G4={ov.cell(4, 7).value} ref_Vehicle G3? F6/G6={ov.cell(3, 6).value}/{ov.cell(3, 7).value}")
    print(f"Dashboard L175 check? Overview J4={ov.cell(4, 10).value}")
    for r in range(1, 293):
        lab = None
        for c in range(2, 12):
            v = ov.cell(r, c).value
            if isinstance(v, str) and v.strip():
                lab = v.strip()
                break
        # get computed label from Computations via Overview C
        c_lab = ov.cell(r, 3).value
        vals = [ov.cell(r, c).value for c in range(17, 25)]
        if any(isinstance(x, (int, float)) and abs(x) > 1e-12 for x in vals if x is not None):
            if isinstance(c_lab, str) or (lab and any(k in str(lab).lower() for k in [
                "ratio", "dscr", "llcr", "check", "irr", "npv", "gear", "plcr"
            ])):
                print(f"R{r}: C={c_lab!r} lab={lab!r} vals2026..={vals}")

    # specifically ratios section - search
    print("\n### Overview all rows with ratio-like labels (data)")
    for r in range(1, 293):
        row_txt = []
        for c in range(1, 16):
            v = ov.cell(r, c).value
            if isinstance(v, str):
                row_txt.append(v)
        joined = " ".join(row_txt).lower()
        if any(k in joined for k in ["ratio", "dscr", "llcr", "plcr", "gear", "irr", "npv", "check b", "check "]):
            vals = [(c, ov.cell(r, c).value) for c in range(1, 30) if ov.cell(r, c).value is not None]
            print(f"R{r}: {vals}")

    wb2.close()
    print("DONE")


if __name__ == "__main__":
    main()
