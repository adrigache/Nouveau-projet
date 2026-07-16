"""One-shot exploration of Holdco / Overview / ass_Vehicle for Phase 5 mapping."""
from __future__ import annotations

import re
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

WB_PATH = "model_real.xlsx"


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip()
    return v


def dump_labels(ws, max_row=None, label_cols=(2, 3, 4, 5, 6, 7, 8, 9)):
    """Print row labels + first formula/value sample for annual cols."""
    max_row = max_row or ws.max_row
    # find first year col: look for 2025/2026 in header rows
    year_col = None
    year_row = None
    for r in range(1, 40):
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(r, c).value
            if v in (2025, 2026, 2027) or (isinstance(v, str) and v.strip() in ("2025", "2026")):
                year_col = c
                year_row = r
                break
        if year_col:
            break
    print(f"  year_header row={year_row} first_year_col={year_col} ({get_column_letter(year_col) if year_col else '?'})")
    # find 2026 col
    col_2026 = None
    col_2027 = None
    if year_row:
        for c in range(1, min(ws.max_column, 120) + 1):
            v = ws.cell(year_row, c).value
            if v == 2026 or (isinstance(v, str) and "2026" in str(v)):
                col_2026 = c
            if v == 2027 or (isinstance(v, str) and "2027" in str(v)):
                col_2027 = c
    print(f"  col_2026={col_2026} ({get_column_letter(col_2026) if col_2026 else '?'}) "
          f"col_2027={col_2027} ({get_column_letter(col_2027) if col_2027 else '?'})")

    for r in range(1, max_row + 1):
        labels = []
        for c in label_cols:
            v = cell_str(ws.cell(r, c).value)
            if v is not None and v != "":
                labels.append(f"C{c}:{v}")
        if not labels:
            continue
        sample = None
        for c in (col_2026, col_2027, year_col, 17, 18, 19):
            if c is None:
                continue
            v = ws.cell(r, c).value
            if v is not None:
                sample = f"{get_column_letter(c)}={v}"
                break
        print(f"R{r:4d} | {' | '.join(labels)[:120]} || {sample}")


def section_keywords(ws, keywords, max_row=None):
    max_row = max_row or min(ws.max_row, 1300)
    hits = []
    for r in range(1, max_row + 1):
        for c in range(1, min(12, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                low = v.lower()
                for kw in keywords:
                    if kw.lower() in low:
                        hits.append((r, c, v))
                        break
    return hits


def main():
    print("Loading formulas workbook...")
    wb = openpyxl.load_workbook(WB_PATH, data_only=False)
    print("Sheets:", wb.sheetnames)

    # ========== Holdco_1 labels ==========
    print("\n" + "=" * 80)
    print("HOLDCO_1 KEYWORD HITS")
    print("=" * 80)
    ws = wb["Holdco_1"]
    kws = [
        "cash flow", "CFADS", "waterfall", "P&L", "profit", "balance", "B/S",
        "senior", "CAPEX", "depreciation", "amort", "CIT", "corporate",
        "share capital", "retained", "equity", "SHL", "cash pooling",
        "detention", "dividend", "EBITDA", "EBIT", "agency", "drawdown",
        "repayment", "interest", "NBV", "assets", "liabilit", "proceeds",
        "Check", "ratio", "DSRF", "junior", "legal reserve",
    ]
    hits = section_keywords(ws, kws)
    for r, c, v in hits:
        print(f"  R{r} C{c}: {v[:100]}")

    print("\n" + "=" * 80)
    print("HOLDCO_1 FULL LABEL DUMP (rows with labels)")
    print("=" * 80)
    dump_labels(ws, max_row=min(ws.max_row, 700))

    print("\n" + "=" * 80)
    print("HOLDCO_2 KEYWORD HITS (compare structure)")
    print("=" * 80)
    ws2 = wb["Holdco_2"]
    hits2 = section_keywords(ws2, ["senior", "detention", "share capital", "CAPEX", "CFADS", "Check"])
    for r, c, v in hits2[:80]:
        print(f"  R{r} C{c}: {v[:100]}")

    print("\n" + "=" * 80)
    print("OVERVIEW KEYWORD HITS")
    print("=" * 80)
    ov = wb["Overview"]
    print(f"Overview max_row={ov.max_row} max_col={ov.max_column}")
    for r in range(1, min(80, ov.max_row) + 1):
        vals = [(c, ov.cell(r, c).value) for c in range(1, min(20, ov.max_column) + 1)
                if ov.cell(r, c).value is not None]
        if vals:
            print(f"R{r}: {vals}")

    hits_ov = section_keywords(ov, [
        "Check", "ratio", "DSCR", "LLCR", "PLCR", "gearing", "entity",
        "vehicle", "Holdco", "SPV", "B/S", "balance", "select", "IRR",
        "NPV", "senior", "equity", "CFADS",
    ])
    print("\nOverview keyword hits:")
    for r, c, v in hits_ov:
        print(f"  R{r} C{c}: {v[:120]}")

    print("\n" + "=" * 80)
    print("OVERVIEW LABEL DUMP")
    print("=" * 80)
    dump_labels(ov, max_row=min(ov.max_row, 400), label_cols=(1, 2, 3, 4, 5, 6, 7, 8, 9, 10))

    # ========== ass_Vehicle senior / financing rows ==========
    print("\n" + "=" * 80)
    print("ASS_VEHICLE financing-related rows (cols M-U = Topco..SPV_6)")
    print("=" * 80)
    veh = wb["ass_Vehicle"]
    # print row labels col B/C/D and values for Holdco_1 (col O typically?)
    # VEH_COLS = 12..20 = M..U
    header = [(c, veh.cell(2, c).value) for c in range(12, 21)]
    print("Vehicle header row2:", header)

    for r in range(1, min(veh.max_row, 500) + 1):
        lab = None
        for c in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11):
            v = veh.cell(r, c).value
            if isinstance(v, str) and v.strip():
                lab = v.strip()
                break
        if lab is None:
            continue
        low = lab.lower()
        if any(k in low for k in [
            "senior", "debt", "facility", "interest", "margin", "swap",
            "capex", "drawdown", "repay", "tenor", "amort", "dsrf",
            "junior", "crowd", "agency", "upfront", "commitment",
            "gear", "ltv", "funding", "financing", "holdco", "topco",
            "share capital", "shl", "cash pool", "detention", "held",
            "vehicle", "perimeter", "opex", "fee",
        ]):
            vals = {veh.cell(2, c).value: veh.cell(r, c).value for c in range(12, 21)}
            print(f"R{r:4d}: {lab[:80]} -> {vals}")

    print("\n" + "=" * 80)
    print("ASS_MODEL senior / financing params")
    print("=" * 80)
    am = wb["ass_Model"]
    for r in range(1, min(am.max_row, 200) + 1):
        lab = None
        for c in (2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
            v = am.cell(r, c).value
            if isinstance(v, str) and v.strip():
                lab = v.strip()
                break
        if lab is None:
            continue
        low = lab.lower()
        if any(k in low for k in [
            "senior", "debt", "facility", "interest", "margin", "dsrf",
            "junior", "agency", "upfront", "commitment", "gear", "ltv",
            "financing", "tenor", "amort", "swap", "euribor", "rate",
        ]):
            # print nearby numeric cols
            vals = [(c, am.cell(r, c).value) for c in range(12, 30)
                    if am.cell(r, c).value is not None]
            print(f"R{r:4d}: {lab[:90]} | {vals[:8]}")

    wb.close()

    # ========== data_only for YE2026 values ==========
    print("\n" + "=" * 80)
    print("Loading DATA_ONLY for YE2026 Holdco_1 B/S & key CF...")
    print("=" * 80)
    wb2 = openpyxl.load_workbook(WB_PATH, data_only=True)
    ws = wb2["Holdco_1"]
    # re-find 2026 col
    year_row = None
    col_2026 = None
    for r in range(1, 40):
        for c in range(1, 80):
            if ws.cell(r, c).value == 2026:
                year_row, col_2026 = r, c
                break
        if col_2026:
            break
    print(f"year_row={year_row} col_2026={col_2026}")

    # dump all non-empty values in col_2026 for labeled rows
    for r in range(1, min(ws.max_row, 700) + 1):
        labels = []
        for c in range(2, 10):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labels.append(v.strip())
        val = ws.cell(r, col_2026).value if col_2026 else None
        if labels and val is not None and val != 0:
            print(f"R{r:4d} YE2026={val!r:20} | {' / '.join(labels)[:100]}")

    print("\n--- SPV_6 CF key rows 2026/2027 ---")
    spv6 = wb2["SPV_6"]
    yr = None
    c26 = c27 = None
    for r in range(1, 40):
        for c in range(1, 80):
            if spv6.cell(r, c).value == 2026:
                yr, c26 = r, c
            if spv6.cell(r, c).value == 2027:
                c27 = c
        if c26 and c27:
            break
    print(f"SPV_6 year_row={yr} c26={c26} c27={c27}")
    for r in range(1, min(spv6.max_row, 600) + 1):
        labels = []
        for c in range(2, 12):
            v = spv6.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labels.append(v.strip())
        if not labels:
            continue
        lab = " / ".join(labels)
        low = lab.lower()
        if any(k in low for k in [
            "cfads", "cash available", "shareholder", "proceeds", "ebitda",
            "cit", "corporate", "cash pooling", "dividend", "shl", "wcr",
            "operating", "capex", "last actual",
        ]):
            v26 = spv6.cell(r, c26).value if c26 else None
            v27 = spv6.cell(r, c27).value if c27 else None
            print(f"R{r:4d} 2026={v26!r} 2027={v27!r} | {lab[:110]}")

    print("\n--- Overview YE2026 non-zero ratios / checks ---")
    ov = wb2["Overview"]
    yr = c26 = None
    for r in range(1, 40):
        for c in range(1, 40):
            if ov.cell(r, c).value == 2026:
                yr, c26 = r, c
                break
        if c26:
            break
    print(f"Overview year_row={yr} c26={c26}")
    for r in range(1, min(ov.max_row, 500) + 1):
        labels = []
        for c in range(1, 12):
            v = ov.cell(r, c).value
            if isinstance(v, str) and v.strip():
                labels.append(v.strip())
        if not labels:
            continue
        vals = []
        if c26:
            for dc in range(0, 8):
                vv = ov.cell(r, c26 + dc).value
                if vv is not None:
                    vals.append((2026 + dc, vv))
        lab = " / ".join(labels)
        low = lab.lower()
        if any(k in low for k in [
            "check", "ratio", "dscr", "llcr", "plcr", "gear", "irr",
            "npv", "bs", "b/s", "entity", "vehicle", "holdco", "spv",
            "selected", "senior", "cfads", "equity", "debt",
        ]) or (vals and any(isinstance(v, (int, float)) and abs(v) > 1e-9 for _, v in vals)):
            # only print interesting
            if any(k in low for k in [
                "check", "ratio", "dscr", "llcr", "plcr", "gear", "irr",
                "npv", "b/s", "entity", "vehicle", "selected", "senior",
                "holdco", "spv", "cfads",
            ]):
                print(f"R{r:4d} {vals[:6]} | {lab[:120]}")

    wb2.close()
    print("\nDONE")


if __name__ == "__main__":
    main()
