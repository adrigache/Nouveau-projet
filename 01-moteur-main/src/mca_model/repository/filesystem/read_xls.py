import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
    
from mca_model.config import (
    rc,
    FNAME_MODEL,
    MODEL_SHEETS, 
    MODEL_SHEETS_COMPUTATIONS)

from mca_model.plumbing import build

from . import xls_dashboard
from . import xls_model
from . import xls_vehicle
from . import xls_asset
from . import xls_market


def rename_first_column(df:pd.DataFrame, name:str='__code'):
    """"""
    my = df.rename(columns={df.columns[0]:name})

    # name unammed
    
    mask = pd.isnull(my.__code)
    my.__code = my.__code.astype(str)
    my.loc[mask,'__code'] = [ f'__unammed_{i:03}' for i in range(mask[mask].index.size)]

    return my.set_index(name, verify_integrity=True)


def _read_excel(a:Path, b:str, silent:bool=False):
    """helper"""
    if not silent:
        rc.wait(f'reading model: {a.name}.{b}')

    my = pd.read_excel(a, sheet_name=b)
    
    if not silent:
        rc.clear_last()
        
    my = rename_first_column(my)
    return my


def _read_excel_raw(a:Path, sheets:dict, silent:bool=False):
    """helper"""

    if not silent:
        rc.wait(f'reading raw model: {a.name}')

    wb = load_workbook(a, read_only=True, data_only=False)

    if not silent:
        rc.clear_last()

    def process(k:str, w:str):
        data = w.iter_rows(min_col=1, max_col=20, values_only=True)
        df = pd.DataFrame(data) 
        if not silent:
            rc.shift(f"sheet '{k}' loaded ... {df.index.size} x {df.columns.size}", 1)
        return rename_first_column(df)

    return {k: process(k, wb[v]) for k,v in sheets.items()}

    
def load_full_model(fname:Path=FNAME_MODEL, silent:bool=False, raw:bool=False):
    """"""
    
    model = {}
    if raw:
        sheets = {**MODEL_SHEETS, **MODEL_SHEETS_COMPUTATIONS}
        model = _read_excel_raw(fname, sheets, silent)
    else:
        for k, sheet in MODEL_SHEETS.items():
            model[k] = _read_excel(fname, sheet, silent)
        
    return model


def postscriptum(m:dict):
    """some modifications"""

    m['dashboard']['scenario'] = {1:"lender", 2:'sponsor'}[m['dashboard']['scenario']]
    m['model']['t_freq'] =  'ME' # month end
        



    
def _load_model(src:Path|dict[str,pd.DataFrame], debug:bool=False, return_field_name:bool=False):
    """read all values from model file"""
    
    model = {}
    modules = {
        'dashboard':xls_dashboard,
        'model':xls_model,
        'market':xls_market,
        'vehicle':xls_vehicle,
        'asset':xls_asset,
        }

    params = dict(debug=debug, return_field_name=return_field_name)

    for k, m in modules.items():

        if isinstance(src, Path):
            if debug:
                rc.check(f'loading model from {src.name}.{k}')
            raw = pd.read_excel(src, MODEL_SHEETS[k])
            raw = rename_first_column(raw)
        else:
            raw = src[k]
            
        # -- if raw return
        params['return_field_name_offset'] = 1 if k == 'dashboard' else 0
        
        # extract raw data
        out = m.load(raw, model, **params)
        if out is not None:
            model[k] = out
    #

    if return_field_name:
        return model
        
    # adapt
    postscriptum(model)   
    return model


        
def load_model(src:Path|dict[str,pd.DataFrame], debug:bool=False, output:str='dict'):
    """emtry point"""

    assert(output in ['dict', 'raw', 'model'])
    if output == 'raw':
        return _load_model(src, debug, return_field_name=True)

    my = _load_model(src, debug, return_field_name=False)
    if output == 'dict':
        return my
    if output == 'model':
        return build.make_model(my, debug=debug)


    
