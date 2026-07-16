"""
One-shot : extrait les echeanciers Holdco -> test/assets/holdco_schedules.toml

INPUTS de plan de dette / postes propres (comme capex_schedule).
Runtime = TOML → moteur. Regenerer apres changement Excel :

  .venv\\Scripts\\python.exe _extract_holdco_schedules.py

``senior_schedule_mode = imported`` : le sculpting DSCR n'est pas encore
moteur ; les series senior/EBL/fees/RE propres sont lues depuis ce fichier.
Les detentions (SC/SHL/CP enfants) restent calculees par le moteur.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import tomli_w

SRC = Path('model_real.xlsx')
OUT = Path('test/assets/holdco_schedules.toml')
YEAR_MAX = 2040


def main():
    print('loading', SRC)
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    doc = {
        '_meta': {
            'source': str(SRC),
            'senior_schedule_mode': 'imported',
            'note': 'Detentions calculees ; postes propres / senior importes.',
        }
    }
    for name in ('Holdco_1', 'Holdco_2'):
        ws = wb[name]
        cols = []
        for c in range(15, 55):
            v = ws.cell(3, c).value
            if getattr(v, 'year', None) and int(v.year) <= YEAR_MAX:
                cols.append((c, int(v.year)))

        def series(row):
            return {str(y): float(ws.cell(row, c).value or 0.0) for c, y in cols}

        repay = {str(y): 0.0 for _, y in cols}
        for r in range(643, 651):
            for c, y in cols:
                repay[str(y)] += float(ws.cell(r, c).value or 0.0)

        doc[name] = {
            'senior_facility_activation': True,
            'senior_schedule_mode': 'imported',
            'senior_drawdown_annual': series(83),
            'senior_repayment_annual': repay,
            'senior_interest_annual': series(144),  # P&L total senior interests
            'senior_eop_annual': series(651),
            'ebl_eop_annual': series(177),
            'vat_facility_eop_annual': series(179),
            'fee_add_annual': series(403),
            'da_fees_annual': series(429),
            'agency_fees_annual': series(66),
            'own_share_capital_annual': series(169),
            'own_legal_reserves_annual': series(170),
            'own_retained_earnings_annual': series(171),
            'own_shl_annual': series(173),
            'own_cit_payable_annual': series(180),
            'own_proceeds_annual': series(125),
            'own_nbv_annual': series(430),
            'own_trade_receivables_annual': series(165),
            'own_cfads_annual': series(91),
        }
        print(name, 'senior2026=', doc[name]['senior_eop_annual'].get('2026'))
    wb.close()
    with open(OUT, 'wb') as f:
        tomli_w.dump(doc, f)
    print('->', OUT)


if __name__ == '__main__':
    main()
